import pandas as pd
import subprocess
import re
import os
import math
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

def process_images_and_extract_data(excel_path, image_dir, proc_script_path, output_excel_path, intermediate_excel_path):
    # Lire le fichier Excel
    df = pd.read_excel(excel_path)

    # Initialiser le compteur d'images traitées et la liste des images traitées
    images_processed = 0
    processed_images = set()
    data_list = []

    def extract_data(text, bridge_df):
        data = {}
        try:
            tif_image_name_match = re.search(r'tif_image_name\s+([^\s]+)', text)
            alpha_left_match = re.search(r'alpha left=([0-9.]+)', text)
            alpha_right_match = re.search(r'alpha right=([0-9.]+)', text)
            yl_match = re.search(r'yl=([0-9.]+) mm', text)
            yc_match = re.search(r'yc=([0-9.]+) mm', text)
            theta_cone_left_match = re.search(r'theta cone left=([0-9.]+)', text)
            theta_cone_right_match = re.search(r'theta cone right=([0-9.]+)', text)
            theta_plate_left_match = re.search(r'theta plate left=([0-9.]+)', text)
            theta_plate_right_match = re.search(r'theta plate right=([0-9.]+)', text)
            
            if tif_image_name_match:
                data['tif_image_name'] = tif_image_name_match.group(1)
            if alpha_left_match:
                data['alpha_left_radian'] = float(alpha_left_match.group(1))
            if alpha_right_match:
                data['alpha_right_radian'] = float(alpha_right_match.group(1))
            if 'alpha_left_radian' in data:
                data['Alpha_left en degré'] = data['alpha_left_radian'] * 180 / 3.14159
            if 'alpha_right_radian' in data:
                data['Alpha_right en degré'] = data['alpha_right_radian'] * 180 / 3.14159
            if yl_match:
                data['yl (mm)'] = float(yl_match.group(1))
            if yc_match:
                data['yc (mm)'] = float(yc_match.group(1))
            if theta_cone_left_match:
                data['theta_cone_left en degré'] = float(theta_cone_left_match.group(1))
            if theta_cone_right_match:
                data['theta_cone_right en degré'] = float(theta_cone_right_match.group(1))
            if theta_plate_left_match:
                data['theta_plate_left'] = float(theta_plate_left_match.group(1))
            if theta_plate_right_match:
                data['theta_plate_right'] = float(theta_plate_right_match.group(1))
            
            if theta_cone_left_match:
                theta_cone_left_degrees = float(theta_cone_left_match.group(1))
                data['theta_cone_left en degré'] = theta_cone_left_degrees
                data['Cos_theta_left'] = math.cos(math.radians(theta_cone_left_degrees))
            if theta_cone_right_match:
                theta_cone_right_degrees = float(theta_cone_right_match.group(1))
                data['theta_cone_right en degré'] = theta_cone_right_degrees
                data['Cos_theta_right'] = math.cos(math.radians(theta_cone_right_degrees))
            if 'yc (mm)' in data:
                data['1/y_c'] = 1 / data['yc (mm)']
            
            # Les valeurs réelles à partir de resultats_bridge.xlsx
            if 'tif_image_name' in data:
                image_name = data['tif_image_name']
                if image_name in bridge_df['tif_image_name'].values:
                    bridge_row = bridge_df.loc[bridge_df['tif_image_name'] == image_name].iloc[0]
                    data['profil_bridge_up (y)'] = bridge_row['profil_bridge_up (y)']
                    data['profil_bridge_dn (y)'] = bridge_row['profil_bridge_dn (y)']
                    data['profil_cone_up (y)'] = bridge_row['profil_cone_up (y)']
                    data['profil_left (x)'] = bridge_row['profil_left (x)']
                    data['profil_right (x)'] = bridge_row['profil_right (x)']
        except (AttributeError, ValueError) as e:
            print(f"Erreur d'extraction des données: {e}")
            return None
        
        return data

    # Boucler à travers chaque ligne du dataframe
    for index, row in df.iterrows():
        if images_processed >= 100:
            break
        
        image_name = row['image_name_prefix']
        
        # Vérifier si l'image a déjà été traitée
        if image_name in processed_images:
            print(f"Image {image_name} déjà traitée, passage à l'image suivante.")
            continue
        
        profil_bridge_up = row['profile_bridge_up']
        profil_bridge_dn = row['profile_bridge_dn']
        profil_cone_up = row['profile_cone_up']
        profil_cone_dn = row['profile_cone_dn']
        profil_left = row['profile_all_lf']
        profil_right = row['profile_all_r']
        alpha = row['profile_accuracy_bridge_cone']
        
        # Construire le chemin complet de l'image
        image_path = os.path.join(image_dir, f"{image_name}.tif")
        
        # Vérifier si le fichier image existe avant d'exécuter le script
        if not os.path.isfile(image_path):
            print(f"Image {image_path} non trouvée, passage à l'image suivante.")
            continue
        # Construire la commande à exécuter
        command = [
            "python3", proc_script_path, image_name,
            str(profil_bridge_up), str(profil_bridge_dn),
            str(profil_cone_up), str(profil_cone_dn),
            str(profil_left), str(profil_right), str(alpha)
        ]
        
        try:
            # Exécuter la commande
            result = subprocess.run(command, capture_output=True, text=True, check=True)
            output = result.stdout + result.stderr
            data = extract_data(output, df)  # Passer le DataFrame à la fonction extract_data
            if data:
                data_list.append(data)
        except subprocess.CalledProcessError as e:
            print(f"Erreur lors de l'exécution de la commande pour l'image {image_name}: {e}")
            continue
        
        # Incrémenter le compteur d'images traitées et ajouter l'image à la liste des images traitées
        images_processed += 1
        processed_images.add(image_name)

        # Générer le fichier Excel intermédiaire après 12 images traitées
        if images_processed == 12:
            print("Génération du fichier Excel intermédiaire après 12 images traitées.")
            df_intermediate = pd.DataFrame(data_list)
            df_intermediate.to_excel(intermediate_excel_path, index=False)
            print(f"Les résultats intermédiaires ont été exportés vers {intermediate_excel_path}")

    print("Traitement des images terminé.")

    # Création du DataFrame final avec l'ordre des colonnes spécifié
    columns_order = [
        'tif_image_name', 'profil_bridge_up (y)', 'profil_bridge_dn (y)', 'profil_cone_up (y)', 'profil_bridge_dn (y)', 
        'profil_left (x)', 'profil_right (x)', 'alpha_left_radian', 'alpha_right_radian', 'Alpha_right en degré', 
        'Alpha_left en degré', 'yl (mm)', 'yc (mm)', 'theta_cone_left en degré', 'theta_cone_right en degré', 
        'theta_plate_left', 'theta_plate_right', 'Cos_theta_left', 'Cos_theta_right', '1/y_c'
    ]

    df_final = pd.DataFrame(data_list, columns=columns_order)

    # Vérifier le contenu du DataFrame avant de l'exporter
    print("DataFrame content:")
    print(df_final.head())

    # Exportation vers un fichier Excel final
    df_final.to_excel(output_excel_path, index=False)

    print(f"Les résultats finaux ont été exportés vers {output_excel_path}")
    print("Traitement terminé.")
    
    # Ajouter les graphiques
    add_charts_to_excel(output_excel_path, df_final)

def add_charts_to_excel(file_path, df):
    wb = load_workbook(file_path)
    ws = wb.active

    # Trouver les indices des colonnes par leurs noms
    headers = {cell.value: cell.column for cell in ws[1]}
    col_yc = headers['yc (mm)']
    col_theta_cone_right = headers['theta_cone_right en degré']
    col_cos_theta_right = headers['Cos_theta_right']
    col_one_over_yc = headers['1/y_c']
    col_theta_cone_left = headers['theta_cone_left en degré']
    col_cos_theta_left = headers['Cos_theta_left']

    # Créer le premier graphique : theta_cone_right en degré vs yc (mm)
    chart1 = ScatterChart()
    chart1.title = "Theta Cone Right (degré) vs Yc (mm)"
    chart1.x_axis.title = "Yc (mm)"
    chart1.y_axis.title = "Theta Cone Right (degré)"
    chart1.style = 13
    chart1.x_axis.majorGridlines = None
    chart1.y_axis.majorGridlines = None
    chart1.legend = None
    
    xvalues1 = Reference(ws, min_col=col_yc, min_row=2, max_row=ws.max_row)
    yvalues1 = Reference(ws, min_col=col_theta_cone_right, min_row=2, max_row=ws.max_row)
    series1 = Series(yvalues1, xvalues1, title_from_data=True)
    series1.marker.symbol = "circle"
    series1.graphicalProperties.line.noFill = True
    series1.marker.graphicalProperties.solidFill = "0000FF"  # Points en bleu
    chart1.series.append(series1)
    ws.add_chart(chart1, "A20")  # Positionner le graphique dans la feuille de calcul

    # Créer le deuxième graphique : Cos_theta_right vs 1/y_c
    chart2 = ScatterChart()
    chart2.title = "Cos Theta Right vs 1/Yc"
    chart2.x_axis.title = "1/Yc"
    chart2.y_axis.title = "Cos Theta Right"
    chart2.style = 13
    chart2.x_axis.majorGridlines = None
    chart2.y_axis.majorGridlines = None
    chart2.legend = None
    
    xvalues2 = Reference(ws, min_col=col_one_over_yc, min_row=2, max_row=ws.max_row)
    yvalues2 = Reference(ws, min_col=col_cos_theta_right, min_row=2, max_row=ws.max_row)
    series2 = Series(yvalues2, xvalues2, title_from_data=True)
    series2.marker.symbol = "circle"
    series2.graphicalProperties.line.noFill = True
    series2.marker.graphicalProperties.solidFill = "0000FF"  # Points en bleu
    chart2.series.append(series2)
    ws.add_chart(chart2, "A35")  # Positionner le graphique dans la feuille de calcul

    # Créer le troisième graphique : Cos_theta_left vs 1/y_c
    chart3 = ScatterChart()
    chart3.title = "Cos Theta Left vs 1/Yc"
    chart3.x_axis.title = "1/Yc"
    chart3.y_axis.title = "Cos Theta Left"
    chart3.style = 13
    chart3.x_axis.majorGridlines = None
    chart3.y_axis.majorGridlines = None
    chart3.legend = None
    
    xvalues3 = Reference(ws, min_col=col_one_over_yc, min_row=2, max_row=ws.max_row)
    yvalues3 = Reference(ws, min_col=col_cos_theta_left, min_row=2, max_row=ws.max_row)
    series3 = Series(yvalues3, xvalues3, title_from_data=True)
    series3.marker.symbol = "circle"
    series3.graphicalProperties.line.noFill = True
    series3.marker.graphicalProperties.solidFill = "0000FF"  # Points en bleu
    chart3.series.append(series3)
    ws.add_chart(chart3, "A50")  # Positionner le graphique dans la feuille de calcul

    # Créer le quatrième graphique : theta_cone_left en degré vs yc (mm)
    chart4 = ScatterChart()
    chart4.title = "Theta Cone Left (degré) vs Yc (mm)"
    chart4.x_axis.title = "Yc (mm)"
    chart4.y_axis.title = "Theta Cone Left (degré)"
    chart4.style = 13
    chart4.x_axis.majorGridlines = None
    chart4.y_axis.majorGridlines = None
    chart4.legend = None
    
    xvalues4 = Reference(ws, min_col=col_yc, min_row=2, max_row=ws.max_row)
    yvalues4 = Reference(ws, min_col=col_theta_cone_left, min_row=2, max_row=ws.max_row)
    series4 = Series(yvalues4, xvalues4, title_from_data=True)
    series4.marker.symbol = "circle"
    series4.graphicalProperties.line.noFill = True
    series4.marker.graphicalProperties.solidFill = "0000FF"  # Points en bleu
    chart4.series.append(series4)
    ws.add_chart(chart4, "A65")  # Positionner le graphique dans la feuille de calcul

    # Sauvegarder le fichier Excel avec les graphiques
    wb.save(file_path)
    print(f"Graphiques ajoutés et fichier sauvegardé sous {file_path}")

# Appel de la fonction principale avec les chemins appropriés
process_images_and_extract_data(
    excel_path="/media/dabaghia/UBUNTU 24_0/Mes doc Windows/Stage2024/antoine_a_traiter/12_19_plan_cone_inf/alpha_5/2/resultats_bridge.xlsx",
    image_dir="/media/dabaghia/UBUNTU 24_0/Mes doc Windows/Stage2024/antoine_a_traiter/12_19_plan_cone_inf/alpha_5/2/",
    proc_script_path="proc_plate_cone.py",
    output_excel_path="/media/dabaghia/UBUNTU 24_0/Mes doc Windows/Stage2024/antoine_a_traiter/12_19_plan_cone_inf/alpha_5/2/resultats_excel.xlsx",
    intermediate_excel_path="/media/dabaghia/UBUNTU 24_0/Mes doc Windows/Stage2024/antoine_a_traiter/12_19_plan_cone_inf/alpha_5/2/resultats_intermediaires.xlsx"
)
